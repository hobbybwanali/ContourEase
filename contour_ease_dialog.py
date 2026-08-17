# -*- coding: utf-8 -*-
"""
ContourEase Dialog - Main GUI and processing logic
"""

import os
import tempfile
import traceback

from qgis.PyQt import uic
from qgis.PyQt.QtCore import Qt, QVariant, QCoreApplication
from qgis.PyQt.QtGui import QColor, QFont
from qgis.PyQt.QtWidgets import (
    QDialog, QFileDialog, QMessageBox, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QPushButton, QComboBox, QCheckBox, QSpinBox,
    QDoubleSpinBox, QGroupBox, QFormLayout, QProgressBar, QTextEdit,
    QRadioButton, QButtonGroup, QWidget, QTabWidget, QFrame
)

from qgis.core import (
    QgsProject, QgsVectorLayer, QgsRasterLayer, QgsField, QgsFeature,
    QgsGeometry, QgsPointXY, QgsCoordinateReferenceSystem,
    QgsCoordinateTransform, QgsCoordinateTransformContext,
    QgsVectorFileWriter, QgsWkbTypes, QgsSymbol, QgsLineSymbol,
    QgsSingleSymbolRenderer, QgsRuleBasedRenderer, QgsPalLayerSettings,
    QgsTextFormat, QgsVectorLayerSimpleLabeling, QgsProperty,
    QgsGraduatedSymbolRenderer, QgsRendererRange, QgsColorRampShader,
    QgsRasterShader, QgsSingleBandPseudoColorRenderer, QgsStyle,
    QgsProcessingFeedback, QgsApplication, QgsMapLayerProxyModel,
    QgsMarkerSymbol, QgsUnitTypes
)
from qgis.gui import QgsProjectionSelectionWidget, QgsFileWidget

import processing


class ContourEaseDialog(QDialog):
    """Dialog for ContourEase plugin."""

    def __init__(self, iface, parent=None):
        super().__init__(parent)
        self.iface = iface
        self.setWindowTitle("ContourEase - Create Professional Contour Maps")
        self.setMinimumWidth(620)
        self.setMinimumHeight(700)
        self.resize(680, 780)

        self.input_path = None
        self.points_layer = None
        self.dem_layer = None
        self.contour_layer = None
        self.boundary_layer = None
        self.temp_dir = tempfile.mkdtemp(prefix="contourease_")

        self._build_ui()
        self._connect_signals()

    def _build_ui(self):
        main_layout = QVBoxLayout(self)

        # Title
        title = QLabel("<h2>ContourEase</h2>")
        title.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(title)

        subtitle = QLabel("Load XYZ data → Transform CRS → Create DEM → Generate Contours")
        subtitle.setAlignment(Qt.AlignCenter)
        subtitle.setStyleSheet("color: #555; margin-bottom: 8px;")
        main_layout.addWidget(subtitle)

        # Tabs
        tabs = QTabWidget()
        main_layout.addWidget(tabs)

        # ========== Tab 1: Data ==========
        data_tab = QWidget()
        data_layout = QVBoxLayout(data_tab)

        # File selection
        file_group = QGroupBox("1. Input XYZ Data (CSV or Excel)")
        file_form = QFormLayout(file_group)

        self.file_edit = QLineEdit()
        self.file_edit.setPlaceholderText("Select a .csv or .xlsx file with X, Y, Z columns...")
        self.btn_browse = QPushButton("Browse...")
        file_row = QHBoxLayout()
        file_row.addWidget(self.file_edit)
        file_row.addWidget(self.btn_browse)
        file_form.addRow("Data file:", file_row)

        self.cmb_x = QComboBox()
        self.cmb_y = QComboBox()
        self.cmb_z = QComboBox()
        file_form.addRow("X / Easting / Longitude column:", self.cmb_x)
        file_form.addRow("Y / Northing / Latitude column:", self.cmb_y)
        file_form.addRow("Z / Elevation / Value column:", self.cmb_z)

        self.chk_header = QCheckBox("First row is header")
        self.chk_header.setChecked(True)
        file_form.addRow(self.chk_header)

        data_layout.addWidget(file_group)

        # CRS section
        crs_group = QGroupBox("2. Coordinate Reference System")
        crs_form = QFormLayout(crs_group)

        self.src_crs = QgsProjectionSelectionWidget()
        self.src_crs.setCrs(QgsCoordinateReferenceSystem("EPSG:4326"))
        crs_form.addRow("Source CRS (of the file):", self.src_crs)

        self.chk_transform = QCheckBox("Transform coordinates to a different CRS")
        self.chk_transform.setChecked(False)
        crs_form.addRow(self.chk_transform)

        self.tgt_crs = QgsProjectionSelectionWidget()
        self.tgt_crs.setCrs(QgsCoordinateReferenceSystem("EPSG:32633"))  # example UTM
        self.tgt_crs.setEnabled(False)
        crs_form.addRow("Target CRS (e.g. UTM):", self.tgt_crs)

        self.lbl_utm_hint = QLabel(
            "<i>Tip: For Lat/Long → UTM, set Source = EPSG:4326 and Target = the correct UTM zone "
            "(e.g. EPSG:32632 for UTM 32N).</i>"
        )
        self.lbl_utm_hint.setWordWrap(True)
        crs_form.addRow(self.lbl_utm_hint)

        data_layout.addWidget(crs_group)
        data_layout.addStretch()
        tabs.addTab(data_tab, "Data & CRS")

        # ========== Tab 2: DEM ==========
        dem_tab = QWidget()
        dem_layout = QVBoxLayout(dem_tab)

        dem_group = QGroupBox("3. Digital Elevation Model (Interpolation)")
        dem_form = QFormLayout(dem_group)

        self.cmb_method = QComboBox()
        self.cmb_method.addItems([
            "IDW (Inverse Distance Weighting)",
            "TIN (Triangular Interpolation)"
        ])
        dem_form.addRow("Interpolation method:", self.cmb_method)

        self.spin_cellsize = QDoubleSpinBox()
        self.spin_cellsize.setRange(0.01, 1e7)
        self.spin_cellsize.setDecimals(4)
        self.spin_cellsize.setValue(10.0)
        self.spin_cellsize.setToolTip(
            "Cell size of the output DEM. Smaller = higher resolution but slower. "
            "For geographic CRS use degrees (e.g. 0.0001), for projected use meters."
        )
        dem_form.addRow("Cell size:", self.spin_cellsize)

        self.spin_power = QDoubleSpinBox()
        self.spin_power.setRange(0.1, 10.0)
        self.spin_power.setValue(2.0)
        self.spin_power.setDecimals(1)
        dem_form.addRow("IDW Power:", self.spin_power)

        self.chk_create_dem = QCheckBox("Create and load DEM layer")
        self.chk_create_dem.setChecked(True)
        dem_form.addRow(self.chk_create_dem)

        dem_layout.addWidget(dem_group)

        style_group = QGroupBox("DEM Styling")
        style_form = QFormLayout(style_group)
        self.cmb_dem_ramp = QComboBox()
        self.cmb_dem_ramp.addItems([
            "Terrain (default)",
            "Elevation",
            "Viridis",
            "Spectral",
            "RdYlGn"
        ])
        style_form.addRow("Color ramp:", self.cmb_dem_ramp)
        dem_layout.addWidget(style_group)
        dem_layout.addStretch()
        tabs.addTab(dem_tab, "DEM")

        # ========== Tab 3: Contours ==========
        cont_tab = QWidget()
        cont_layout = QVBoxLayout(cont_tab)

        cont_group = QGroupBox("4. Contour Lines")
        cont_form = QFormLayout(cont_group)

        self.chk_create_contours = QCheckBox("Create contour lines")
        self.chk_create_contours.setChecked(True)
        cont_form.addRow(self.chk_create_contours)

        self.spin_interval = QDoubleSpinBox()
        self.spin_interval.setRange(0.001, 1e6)
        self.spin_interval.setDecimals(3)
        self.spin_interval.setValue(10.0)
        cont_form.addRow("Contour interval:", self.spin_interval)

        self.spin_offset = QDoubleSpinBox()
        self.spin_offset.setRange(-1e6, 1e6)
        self.spin_offset.setDecimals(3)
        self.spin_offset.setValue(0.0)
        cont_form.addRow("Base / Offset:", self.spin_offset)

        self.chk_smooth = QCheckBox("Smooth contours (recommended for clean look)")
        self.chk_smooth.setChecked(True)
        cont_form.addRow(self.chk_smooth)

        self.spin_smooth = QSpinBox()
        self.spin_smooth.setRange(1, 10)
        self.spin_smooth.setValue(2)
        cont_form.addRow("Smoothing iterations:", self.spin_smooth)

        cont_layout.addWidget(cont_group)

        # Boundary / mask
        bound_group = QGroupBox("Boundary / Mask (optional)")
        bound_form = QFormLayout(bound_group)

        self.chk_boundary = QCheckBox("Clip contours to a boundary polygon")
        self.chk_boundary.setChecked(False)
        bound_form.addRow(self.chk_boundary)

        self.boundary_edit = QLineEdit()
        self.boundary_edit.setPlaceholderText("Select a polygon shapefile / GeoPackage / ...")
        self.boundary_edit.setEnabled(False)
        self.btn_boundary = QPushButton("Browse...")
        self.btn_boundary.setEnabled(False)
        b_row = QHBoxLayout()
        b_row.addWidget(self.boundary_edit)
        b_row.addWidget(self.btn_boundary)
        bound_form.addRow("Boundary file:", b_row)

        self.chk_load_boundary = QCheckBox("Also load boundary layer into the project")
        self.chk_load_boundary.setChecked(True)
        self.chk_load_boundary.setEnabled(False)
        bound_form.addRow(self.chk_load_boundary)

        cont_layout.addWidget(bound_group)

        style_c_group = QGroupBox("Contour Styling (professional)")
        style_c_form = QFormLayout(style_c_group)

        self.chk_index = QCheckBox("Highlight index contours (every Nth line thicker)")
        self.chk_index.setChecked(True)
        style_c_form.addRow(self.chk_index)

        self.spin_index_every = QSpinBox()
        self.spin_index_every.setRange(2, 20)
        self.spin_index_every.setValue(5)
        style_c_form.addRow("Index every:", self.spin_index_every)

        self.chk_labels = QCheckBox("Label contours with elevation")
        self.chk_labels.setChecked(True)
        style_c_form.addRow(self.chk_labels)

        self.spin_label_size = QSpinBox()
        self.spin_label_size.setRange(6, 24)
        self.spin_label_size.setValue(9)
        style_c_form.addRow("Label font size:", self.spin_label_size)

        cont_layout.addWidget(style_c_group)
        cont_layout.addStretch()
        tabs.addTab(cont_tab, "Contours")

        # ========== Tab 4: Output ==========
        out_tab = QWidget()
        out_layout = QVBoxLayout(out_tab)

        out_group = QGroupBox("5. Output Options")
        out_form = QFormLayout(out_group)

        self.chk_add_points = QCheckBox("Also load original points as a layer")
        self.chk_add_points.setChecked(True)
        out_form.addRow(self.chk_add_points)

        self.chk_save = QCheckBox("Save results to disk (otherwise temporary layers)")
        self.chk_save.setChecked(False)
        out_form.addRow(self.chk_save)

        self.out_dir_edit = QLineEdit()
        self.out_dir_edit.setPlaceholderText("Output folder...")
        self.out_dir_edit.setEnabled(False)
        self.btn_out_dir = QPushButton("Browse...")
        self.btn_out_dir.setEnabled(False)
        out_row = QHBoxLayout()
        out_row.addWidget(self.out_dir_edit)
        out_row.addWidget(self.btn_out_dir)
        out_form.addRow("Output folder:", out_row)

        out_layout.addWidget(out_group)

        # DXF export
        dxf_group = QGroupBox("6. Export to DXF (optional)")
        dxf_form = QFormLayout(dxf_group)

        self.chk_dxf = QCheckBox("Export selected layers to DXF after processing")
        self.chk_dxf.setChecked(False)
        dxf_form.addRow(self.chk_dxf)

        self.chk_dxf_contours = QCheckBox("Contours (with ELEV attribute)")
        self.chk_dxf_contours.setChecked(True)
        self.chk_dxf_contours.setEnabled(False)
        dxf_form.addRow(self.chk_dxf_contours)

        self.chk_dxf_boundary = QCheckBox("Boundary polygon")
        self.chk_dxf_boundary.setChecked(True)
        self.chk_dxf_boundary.setEnabled(False)
        dxf_form.addRow(self.chk_dxf_boundary)

        self.chk_dxf_points = QCheckBox("XYZ points")
        self.chk_dxf_points.setChecked(False)
        self.chk_dxf_points.setEnabled(False)
        dxf_form.addRow(self.chk_dxf_points)

        self.dxf_path_edit = QLineEdit()
        self.dxf_path_edit.setPlaceholderText("Output .dxf file path...")
        self.dxf_path_edit.setEnabled(False)
        self.btn_dxf_path = QPushButton("Browse...")
        self.btn_dxf_path.setEnabled(False)
        dxf_row = QHBoxLayout()
        dxf_row.addWidget(self.dxf_path_edit)
        dxf_row.addWidget(self.btn_dxf_path)
        dxf_form.addRow("DXF file:", dxf_row)

        out_layout.addWidget(dxf_group)

        # Log
        log_group = QGroupBox("Log")
        log_layout = QVBoxLayout(log_group)
        self.log = QTextEdit()
        self.log.setReadOnly(True)
        self.log.setMaximumHeight(160)
        log_layout.addWidget(self.log)
        out_layout.addWidget(log_group)

        self.progress = QProgressBar()
        self.progress.setValue(0)
        out_layout.addWidget(self.progress)

        out_layout.addStretch()
        tabs.addTab(out_tab, "Output & Run")

        # Buttons
        btn_layout = QHBoxLayout()
        self.btn_run = QPushButton("Create Contours & DEM")
        self.btn_run.setStyleSheet(
            "QPushButton { background-color: #2E7D32; color: white; font-weight: bold; "
            "padding: 10px 20px; border-radius: 4px; }"
            "QPushButton:hover { background-color: #388E3C; }"
            "QPushButton:disabled { background-color: #9E9E9E; }"
        )
        self.btn_close = QPushButton("Close")
        btn_layout.addStretch()
        btn_layout.addWidget(self.btn_run)
        btn_layout.addWidget(self.btn_close)
        main_layout.addLayout(btn_layout)

    def _connect_signals(self):
        self.btn_browse.clicked.connect(self._browse_file)
        self.btn_out_dir.clicked.connect(self._browse_out_dir)
        self.btn_run.clicked.connect(self._run)
        self.btn_close.clicked.connect(self.reject)
        self.chk_transform.toggled.connect(self.tgt_crs.setEnabled)
        self.chk_save.toggled.connect(self._toggle_save)
        self.file_edit.textChanged.connect(self._on_file_changed)
        self.cmb_method.currentIndexChanged.connect(self._on_method_changed)
        self.chk_boundary.toggled.connect(self._toggle_boundary)
        self.btn_boundary.clicked.connect(self._browse_boundary)
        self.chk_dxf.toggled.connect(self._toggle_dxf)
        self.btn_dxf_path.clicked.connect(self._browse_dxf)

    def _toggle_save(self, checked):
        self.out_dir_edit.setEnabled(checked)
        self.btn_out_dir.setEnabled(checked)

    def _toggle_boundary(self, checked):
        self.boundary_edit.setEnabled(checked)
        self.btn_boundary.setEnabled(checked)
        self.chk_load_boundary.setEnabled(checked)

    def _toggle_dxf(self, checked):
        self.chk_dxf_contours.setEnabled(checked)
        self.chk_dxf_boundary.setEnabled(checked)
        self.chk_dxf_points.setEnabled(checked)
        self.dxf_path_edit.setEnabled(checked)
        self.btn_dxf_path.setEnabled(checked)

    def _browse_boundary(self):
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Select boundary polygon",
            "",
            "Vector files (*.shp *.gpkg *.geojson *.json *.kml *.gml);;All files (*)"
        )
        if path:
            self.boundary_edit.setText(path)

    def _browse_dxf(self):
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Save DXF file",
            "contours.dxf",
            "DXF files (*.dxf)"
        )
        if path:
            if not path.lower().endswith(".dxf"):
                path += ".dxf"
            self.dxf_path_edit.setText(path)

    def _on_method_changed(self, idx):
        # Power only relevant for IDW
        self.spin_power.setEnabled(idx == 0)

    def _log(self, msg):
        self.log.append(msg)
        self.log.ensureCursorVisible()
        QCoreApplication.processEvents()

    def _browse_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Select XYZ data file",
            "",
            "Data files (*.csv *.txt *.xlsx *.xls);;CSV (*.csv *.txt);;Excel (*.xlsx *.xls);;All files (*)"
        )
        if path:
            self.file_edit.setText(path)

    def _browse_out_dir(self):
        path = QFileDialog.getExistingDirectory(self, "Select output folder")
        if path:
            self.out_dir_edit.setText(path)

    def _on_file_changed(self, path):
        self.cmb_x.clear()
        self.cmb_y.clear()
        self.cmb_z.clear()
        if not path or not os.path.isfile(path):
            return
        try:
            cols = self._read_columns(path)
            if cols:
                self.cmb_x.addItems(cols)
                self.cmb_y.addItems(cols)
                self.cmb_z.addItems(cols)
                # Heuristic auto-select
                lower = [c.lower() for c in cols]
                for i, c in enumerate(lower):
                    if c in ("x", "easting", "east", "lon", "long", "longitude", "lng"):
                        self.cmb_x.setCurrentIndex(i)
                    if c in ("y", "northing", "north", "lat", "latitude"):
                        self.cmb_y.setCurrentIndex(i)
                    if c in ("z", "elev", "elevation", "height", "value", "alt", "altitude", "h"):
                        self.cmb_z.setCurrentIndex(i)
        except Exception as e:
            self._log(f"Could not read columns: {e}")

    def _read_columns(self, path):
        """Return list of column names / indices from CSV or Excel."""
        ext = os.path.splitext(path)[1].lower()
        if ext in (".xlsx", ".xls"):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                ws = wb.active
                rows = ws.iter_rows(min_row=1, max_row=2, values_only=True)
                header = next(rows)
                if self.chk_header.isChecked() and header:
                    return [str(h) if h is not None else f"Col{i+1}" for i, h in enumerate(header)]
                else:
                    # No header – invent names
                    first = header
                    return [f"Col{i+1}" for i in range(len(first))]
            except ImportError:
                # Fallback: try pandas if available
                try:
                    import pandas as pd
                    df = pd.read_excel(path, nrows=1, header=0 if self.chk_header.isChecked() else None)
                    return list(df.columns.astype(str))
                except Exception:
                    raise RuntimeError("Excel support requires openpyxl or pandas. Install with: pip install openpyxl")
        else:
            # CSV / TXT
            import csv
            with open(path, "r", encoding="utf-8-sig", errors="replace") as f:
                # Detect delimiter
                sample = f.read(4096)
                f.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=",;\t ")
                except csv.Error:
                    dialect = csv.excel
                reader = csv.reader(f, dialect)
                first = next(reader)
                if self.chk_header.isChecked():
                    return [str(h).strip() or f"Col{i+1}" for i, h in enumerate(first)]
                else:
                    return [f"Col{i+1}" for i in range(len(first))]

    def _run(self):
        self.progress.setValue(0)
        self.log.clear()
        self.btn_run.setEnabled(False)
        try:
            self._do_process()
        except Exception as e:
            self._log(f"<b style='color:red'>ERROR: {e}</b>")
            self._log(traceback.format_exc())
            QMessageBox.critical(self, "Error", str(e))
        finally:
            self.btn_run.setEnabled(True)
            self.progress.setValue(100)

    def _do_process(self):
        path = self.file_edit.text().strip()
        if not path or not os.path.isfile(path):
            raise ValueError("Please select a valid input file.")

        x_col = self.cmb_x.currentText()
        y_col = self.cmb_y.currentText()
        z_col = self.cmb_z.currentText()
        if not x_col or not y_col or not z_col:
            raise ValueError("Please select X, Y and Z columns.")

        src_crs = self.src_crs.crs()
        if not src_crs.isValid():
            raise ValueError("Source CRS is not valid.")

        do_transform = self.chk_transform.isChecked()
        tgt_crs = self.tgt_crs.crs() if do_transform else src_crs
        if do_transform and not tgt_crs.isValid():
            raise ValueError("Target CRS is not valid.")

        self._log(f"Reading data from: {path}")
        self.progress.setValue(5)

        # 1. Load points into a memory layer
        points = self._load_points(path, x_col, y_col, z_col, src_crs)
        self._log(f"Loaded {points.featureCount()} points.")
        self.progress.setValue(15)

        # 2. Optional CRS transform
        if do_transform and src_crs != tgt_crs:
            self._log(f"Transforming from {src_crs.authid()} → {tgt_crs.authid()} ...")
            points = self._reproject_layer(points, tgt_crs)
            self._log("Reprojection done.")
        else:
            tgt_crs = src_crs

        self.progress.setValue(25)

        if self.chk_add_points.isChecked():
            points.setName("XYZ Points")
            QgsProject.instance().addMapLayer(points)
            self.points_layer = points
            self._style_points(points)

        # 3. Create DEM
        dem = None
        if self.chk_create_dem.isChecked() or self.chk_create_contours.isChecked():
            self._log("Interpolating DEM ...")
            dem = self._create_dem(points, tgt_crs)
            self._log("DEM created.")
            self.progress.setValue(55)

            if self.chk_create_dem.isChecked():
                dem.setName("DEM (interpolated)")
                QgsProject.instance().addMapLayer(dem)
                self.dem_layer = dem
                self._style_dem(dem)

        # 4. Contours
        boundary = None
        if self.chk_boundary.isChecked():
            bpath = self.boundary_edit.text().strip()
            if not bpath or not os.path.isfile(bpath):
                raise ValueError("Boundary clipping is enabled but no valid boundary file was selected.")
            boundary = self._load_boundary(bpath, tgt_crs)
            self.boundary_layer = boundary
            self._log(f"Boundary loaded: {boundary.featureCount()} polygon feature(s).")
            if self.chk_load_boundary.isChecked():
                boundary.setName("Boundary")
                QgsProject.instance().addMapLayer(boundary)

        if self.chk_create_contours.isChecked():
            if dem is None:
                raise ValueError("DEM is required to generate contours.")
            self._log("Generating contours ...")
            contours = self._create_contours(dem)
            self._log(f"Contours created: {contours.featureCount()} features.")
            self.progress.setValue(70)

            if self.chk_smooth.isChecked():
                self._log("Smoothing contours ...")
                contours = self._smooth_contours(contours)
                self._log("Smoothing finished.")

            if boundary is not None:
                self._log("Clipping contours to boundary ...")
                contours = self._clip_to_boundary(contours, boundary)
                self._log(f"Contours after clip: {contours.featureCount()} features.")
                contours = self._ensure_elev_field(contours)

            contours.setName("Contours")
            QgsProject.instance().addMapLayer(contours)
            self.contour_layer = contours
            self._style_contours(contours)
            self.progress.setValue(90)

        # Optional save to folder
        if self.chk_save.isChecked():
            out_dir = self.out_dir_edit.text().strip()
            if out_dir and os.path.isdir(out_dir):
                self._save_results(out_dir, points, dem, self.contour_layer, boundary)
                self._log(f"Results saved to {out_dir}")

        # Optional DXF export
        if self.chk_dxf.isChecked():
            dxf_path = self.dxf_path_edit.text().strip()
            if not dxf_path:
                raise ValueError("DXF export is enabled but no output DXF path was set.")
            self._export_dxf(
                dxf_path,
                contours=self.contour_layer if self.chk_dxf_contours.isChecked() else None,
                boundary=boundary if self.chk_dxf_boundary.isChecked() else None,
                points=points if self.chk_dxf_points.isChecked() else None,
            )
            self._log(f"DXF exported to: {dxf_path}")

        self.progress.setValue(100)
        self._log("<b style='color:green'>Done! Layers added to the project.</b>")
        self.iface.mapCanvas().refresh()
        QMessageBox.information(
            self,
            "Success",
            "Contour map and/or DEM created successfully.\n"
            "Layers have been added to the project."
        )

    def _load_points(self, path, x_col, y_col, z_col, crs):
        """Load XYZ from CSV/Excel into a memory point layer with Z attribute."""
        ext = os.path.splitext(path)[1].lower()
        rows = []

        if ext in (".xlsx", ".xls"):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                ws = wb.active
                data = list(ws.iter_rows(values_only=True))
            except ImportError:
                import pandas as pd
                df = pd.read_excel(path, header=0 if self.chk_header.isChecked() else None)
                data = [list(df.columns)] + df.values.tolist() if self.chk_header.isChecked() else df.values.tolist()
                # Normalize
                if not self.chk_header.isChecked():
                    data = [[f"Col{i+1}" for i in range(len(data[0]))]] + data
            if self.chk_header.isChecked():
                header = [str(h) if h is not None else f"Col{i+1}" for i, h in enumerate(data[0])]
                data_rows = data[1:]
            else:
                header = [f"Col{i+1}" for i in range(len(data[0]))]
                data_rows = data
            try:
                xi = header.index(x_col)
                yi = header.index(y_col)
                zi = header.index(z_col)
            except ValueError:
                raise ValueError(f"Column not found. Available: {header}")
            for r in data_rows:
                if r is None or len(r) <= max(xi, yi, zi):
                    continue
                try:
                    x = float(r[xi])
                    y = float(r[yi])
                    z = float(r[zi])
                    rows.append((x, y, z))
                except (TypeError, ValueError):
                    continue
        else:
            import csv
            with open(path, "r", encoding="utf-8-sig", errors="replace") as f:
                sample = f.read(4096)
                f.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=",;\t ")
                except csv.Error:
                    dialect = csv.excel
                reader = csv.reader(f, dialect)
                first = next(reader)
                if self.chk_header.isChecked():
                    header = [str(h).strip() or f"Col{i+1}" for i, h in enumerate(first)]
                    data_iter = reader
                else:
                    header = [f"Col{i+1}" for i in range(len(first))]
                    data_iter = [first] + list(reader)
                try:
                    xi = header.index(x_col)
                    yi = header.index(y_col)
                    zi = header.index(z_col)
                except ValueError:
                    raise ValueError(f"Column not found. Available: {header}")
                for r in data_iter:
                    if len(r) <= max(xi, yi, zi):
                        continue
                    try:
                        x = float(r[xi].strip())
                        y = float(r[yi].strip())
                        z = float(r[zi].strip())
                        rows.append((x, y, z))
                    except (ValueError, AttributeError):
                        continue

        if not rows:
            raise ValueError("No valid numeric XYZ rows found.")

        vl = QgsVectorLayer(f"Point?crs={crs.authid()}", "xyz_temp", "memory")
        pr = vl.dataProvider()
        pr.addAttributes([
            QgsField("X", QVariant.Double),
            QgsField("Y", QVariant.Double),
            QgsField("Z", QVariant.Double),
        ])
        vl.updateFields()

        feats = []
        for x, y, z in rows:
            f = QgsFeature(vl.fields())
            f.setGeometry(QgsGeometry.fromPointXY(QgsPointXY(x, y)))
            f.setAttributes([x, y, z])
            feats.append(f)
        pr.addFeatures(feats)
        vl.updateExtents()
        return vl

    def _reproject_layer(self, layer, target_crs):
        """Reproject a point layer to target CRS, keeping attributes."""
        result = processing.run(
            "native:reprojectlayer",
            {
                "INPUT": layer,
                "TARGET_CRS": target_crs,
                "OUTPUT": "memory:"
            }
        )
        return result["OUTPUT"]

    def _create_dem(self, points, crs):
        """Create DEM using IDW or TIN."""
        cell = self.spin_cellsize.value()
        method = self.cmb_method.currentIndex()

        # Compute extent from points (with small buffer)
        extent = points.extent()
        w = extent.width()
        h = extent.height()
        buf = max(w, h) * 0.02 if max(w, h) > 0 else cell * 5
        extent = extent.buffered(buf)

        # Write points to a temporary GeoPackage so processing can reliably open them
        tmp_pts = os.path.join(self.temp_dir, "points_for_interp.gpkg")
        QgsVectorFileWriter.writeAsVectorFormat(
            points, tmp_pts, "UTF-8", crs, "GPKG"
        )

        # Re-open and resolve the Z field index from the saved file (most reliable)
        tmp_layer = QgsVectorLayer(tmp_pts, "pts", "ogr")
        z_idx = tmp_layer.fields().indexOf("Z")
        if z_idx < 0:
            # fallback: look for common elevation names
            for name in ("Z", "ELEV", "ELEVATION", "HEIGHT", "VALUE"):
                z_idx = tmp_layer.fields().indexOf(name)
                if z_idx >= 0:
                    break
        if z_idx < 0:
            raise RuntimeError(
                "Could not find Z/elevation field in the point layer. "
                "Make sure you selected the correct Z column."
            )

        # Log Z range so the user can confirm the correct column was used
        try:
            zvals = [float(f["Z"]) for f in points.getFeatures() if f["Z"] is not None]
            if zvals:
                self._log(f"Using Z field (index {z_idx}) range: {min(zvals):.3f} .. {max(zvals):.3f}")
        except (TypeError, ValueError, KeyError) as e:
            self._log(f"Could not log Z range: {e}")

        # INTERPOLATION_DATA format used by QGIS:
        # path::~::use_z(0/1)::~::field_index::~::type(0=Points)
        interp_data = f"{tmp_pts}::~::0::~::{z_idx}::~::0"

        # Extent string: "xmin,xmax,ymin,ymax [CRS]"
        extent_str = (
            f"{extent.xMinimum()},{extent.xMaximum()},"
            f"{extent.yMinimum()},{extent.yMaximum()} [{crs.authid()}]"
        )

        if method == 0:  # IDW
            power = self.spin_power.value()
            params = {
                "INTERPOLATION_DATA": interp_data,
                "DISTANCE_COEFFICIENT": power,
                "EXTENT": extent_str,
                "PIXEL_SIZE": cell,
                "OUTPUT": "TEMPORARY_OUTPUT"
            }
            result = processing.run("qgis:idwinterpolation", params, feedback=None)
            dem = result["OUTPUT"]
        else:  # TIN
            params = {
                "INTERPOLATION_DATA": interp_data,
                "METHOD": 0,  # Linear
                "EXTENT": extent_str,
                "PIXEL_SIZE": cell,
                "OUTPUT": "TEMPORARY_OUTPUT"
            }
            result = processing.run("qgis:tininterpolation", params, feedback=None)
            dem = result["OUTPUT"]

        # Ensure we have a valid QgsRasterLayer
        if isinstance(dem, str):
            dem = QgsRasterLayer(dem, "DEM")
        if not dem.isValid():
            raise RuntimeError(
                "Failed to create a valid DEM raster. "
                "Try a larger cell size or check that your points have valid coordinates."
            )
        dem.setCrs(crs)
        return dem

    def _create_contours(self, dem):
        """Generate contours from DEM using GDAL. Elevation attribute is always named ELEV."""
        interval = self.spin_interval.value()
        offset = self.spin_offset.value()

        # Write DEM to a temp GeoTIFF so gdal:contour always receives a file path
        tmp_dem = os.path.join(self.temp_dir, "dem_for_contours.tif")
        processing.run(
            "gdal:translate",
            {
                "INPUT": dem,
                "TARGET_CRS": None,
                "NODATA": None,
                "COPY_SUBDATASETS": False,
                "OPTIONS": "",
                "EXTRA": "",
                "DATA_TYPE": 0,
                "OUTPUT": tmp_dem,
            },
            feedback=None,
        )

        params = {
            "INPUT": tmp_dem,
            "BAND": 1,
            "INTERVAL": interval,
            "FIELD_NAME": "ELEV",
            "CREATE_3D": False,
            "IGNORE_NODATA": True,
            "NODATA": None,
            "OFFSET": offset,
            "EXTRA": "",
            "OUTPUT": "TEMPORARY_OUTPUT",
        }
        result = processing.run("gdal:contour", params, feedback=None)
        contours = result["OUTPUT"]
        if isinstance(contours, str):
            contours = QgsVectorLayer(contours, "Contours", "ogr")

        # Guarantee an ELEV field exists and holds elevation (Z) values
        contours = self._ensure_elev_field(contours)
        return contours

    def _ensure_elev_field(self, layer):
        """Make sure the layer has a numeric field named ELEV with elevation values."""
        fields = layer.fields()
        names_upper = {f.name().upper(): f.name() for f in fields}

        elev_name = None
        for candidate in ("ELEV", "ELEVATION", "Z", "HEIGHT", "VALUE", "ALTITUDE"):
            if candidate in names_upper:
                elev_name = names_upper[candidate]
                break

        if elev_name is None:
            # Fallback: first numeric field
            for f in fields:
                if f.type() in (QVariant.Double, QVariant.Int, QVariant.LongLong):
                    elev_name = f.name()
                    break

        if elev_name is None:
            raise RuntimeError(
                "Contour layer has no numeric elevation attribute. "
                "Check that the DEM was built from the Z column."
            )

        # If the field is already named ELEV we are done
        if elev_name.upper() == "ELEV":
            return layer

        # Rename / copy into a clean memory layer with field ELEV
        mem = QgsVectorLayer(
            f"LineString?crs={layer.crs().authid()}", "Contours", "memory"
        )
        pr = mem.dataProvider()
        pr.addAttributes([QgsField("ELEV", QVariant.Double)])
        mem.updateFields()

        feats = []
        idx = layer.fields().indexOf(elev_name)
        for f in layer.getFeatures():
            nf = QgsFeature(mem.fields())
            nf.setGeometry(f.geometry())
            try:
                val = float(f.attributes()[idx])
            except (TypeError, ValueError, IndexError):
                continue
            nf.setAttributes([val])
            feats.append(nf)
        pr.addFeatures(feats)
        mem.updateExtents()
        return mem

    def _smooth_contours(self, layer):
        """Apply iterative geometry smoothing while preserving the ELEV attribute."""
        iterations = self.spin_smooth.value()
        current = layer
        for i in range(iterations):
            result = processing.run(
                "native:smoothgeometry",
                {
                    "INPUT": current,
                    "ITERATIONS": 1,
                    "OFFSET": 0.25,
                    "MAX_ANGLE": 180.0,
                    "OUTPUT": "memory:",
                },
            )
            current = result["OUTPUT"]
        # Re-ensure ELEV after smoothing (some versions drop/rename fields)
        current = self._ensure_elev_field(current)
        return current

    def _style_points(self, layer):
        """Simple styling for points."""
        symbol = QgsMarkerSymbol.createSimple({
            "name": "circle",
            "color": "255,80,80,180",
            "outline_color": "120,0,0",
            "outline_width": "0.4",
            "size": "2.5"
        })
        layer.renderer().setSymbol(symbol)
        layer.triggerRepaint()

    def _style_dem(self, layer):
        """Apply a nice color ramp to the DEM."""
        ramp_name = self.cmb_dem_ramp.currentText()
        # Map friendly names to QgsStyle ramp names
        ramp_map = {
            "Terrain (default)": "Terrain",
            "Elevation": "Elevation",
            "Viridis": "Viridis",
            "Spectral": "Spectral",
            "RdYlGn": "RdYlGn"
        }
        style = QgsStyle.defaultStyle()
        ramp = style.colorRamp(ramp_map.get(ramp_name, "Terrain"))
        if ramp is None:
            ramp = style.colorRamp("Viridis")

        stats = layer.dataProvider().bandStatistics(1)
        min_val = stats.minimumValue
        max_val = stats.maximumValue
        if min_val is None or max_val is None or min_val >= max_val:
            min_val, max_val = 0, 100

        shader = QgsRasterShader()
        color_ramp_shader = QgsColorRampShader()
        color_ramp_shader.setColorRampType(QgsColorRampShader.Interpolated)
        color_ramp_shader.setClassificationMode(QgsColorRampShader.Continuous)
        color_ramp_shader.setMinimumValue(min_val)
        color_ramp_shader.setMaximumValue(max_val)

        # Create discrete items from the ramp
        n = 15
        items = []
        for i in range(n + 1):
            val = min_val + (max_val - min_val) * i / n
            color = ramp.color(i / n)
            items.append(QgsColorRampShader.ColorRampItem(val, color, f"{val:.1f}"))
        color_ramp_shader.setColorRampItemList(items)
        shader.setRasterShaderFunction(color_ramp_shader)

        renderer = QgsSingleBandPseudoColorRenderer(layer.dataProvider(), 1, shader)
        layer.setRenderer(renderer)
        layer.triggerRepaint()

    def _style_contours(self, layer):
        """Apply professional styling: index contours thicker, labels from ELEV (Z)."""
        elev_field = "ELEV"
        names = [f.name() for f in layer.fields()]
        if "ELEV" not in names:
            for f in layer.fields():
                if f.name().upper() in ("ELEV", "ELEVATION", "Z", "HEIGHT", "VALUE"):
                    elev_field = f.name()
                    break
            else:
                elev_field = names[0] if names else "ELEV"

        # Log elevation range so user can verify labels use Z, not X/Y
        try:
            vals = []
            for feat in layer.getFeatures():
                v = feat[elev_field]
                if v is not None:
                    vals.append(float(v))
                if len(vals) >= 80:
                    break
            if vals:
                self._log(
                    f"Contour labels use field '{elev_field}' "
                    f"(range {min(vals):.3f} .. {max(vals):.3f})"
                )
        except (TypeError, ValueError, KeyError) as e:
            self._log(f"Could not log contour elevation range: {e}")

        interval = self.spin_interval.value()
        index_every = self.spin_index_every.value() if self.chk_index.isChecked() else 999999

        # Rule-based renderer
        root_rule = QgsRuleBasedRenderer.Rule(None)

        # Index contours (thicker, darker)
        index_symbol = QgsLineSymbol.createSimple({
            "color": "40,40,40",
            "width": "0.7",
            "capstyle": "round",
            "joinstyle": "round"
        })
        index_rule = QgsRuleBasedRenderer.Rule(index_symbol)
        index_rule.setFilterExpression(
            f'remainder("{elev_field}", {interval * index_every}) = 0 OR '
            f'remainder("{elev_field}", {interval * index_every}) < 0.0001'
        )
        index_rule.setLabel("Index contours")
        root_rule.appendChild(index_rule)

        # Intermediate contours
        inter_symbol = QgsLineSymbol.createSimple({
            "color": "80,80,80",
            "width": "0.25",
            "capstyle": "round",
            "joinstyle": "round"
        })
        inter_rule = QgsRuleBasedRenderer.Rule(inter_symbol)
        inter_rule.setFilterExpression("ELSE")
        inter_rule.setLabel("Intermediate contours")
        root_rule.appendChild(inter_rule)

        renderer = QgsRuleBasedRenderer(root_rule)
        layer.setRenderer(renderer)

        # Labels
        if self.chk_labels.isChecked():
            settings = QgsPalLayerSettings()
            settings.fieldName = elev_field
            settings.isExpression = False
            settings.enabled = True
            settings.drawLabels = True
            settings.placement = QgsPalLayerSettings.Line
            settings.placementFlags = QgsPalLayerSettings.OnLine

            text_format = QgsTextFormat()
            text_format.setSize(self.spin_label_size.value())
            text_format.setSizeUnit(QgsUnitTypes.RenderPoints)
            text_format.setColor(QColor(30, 30, 30))
            font = QFont("Arial")
            font.setBold(False)
            text_format.setFont(font)

            # White buffer for readability
            from qgis.core import QgsTextBufferSettings
            buffer = QgsTextBufferSettings()
            buffer.setEnabled(True)
            buffer.setSize(0.8)
            buffer.setColor(QColor(255, 255, 255, 200))
            text_format.setBuffer(buffer)

            settings.setFormat(text_format)

            # Only label index contours if index is enabled
            if self.chk_index.isChecked():
                settings.scaleVisibility = False
                # Filter via data-defined or just label all; for simplicity label all
                # (advanced: use data defined show label)

            labeling = QgsVectorLayerSimpleLabeling(settings)
            layer.setLabeling(labeling)
            layer.setLabelsEnabled(True)

        layer.triggerRepaint()

    def _save_results(self, out_dir, points, dem, contours, boundary=None):
        """Save layers to disk."""
        if points:
            out_pts = os.path.join(out_dir, "xyz_points.gpkg")
            QgsVectorFileWriter.writeAsVectorFormat(
                points, out_pts, "UTF-8", points.crs(), "GPKG"
            )
        if dem and isinstance(dem, QgsRasterLayer):
            out_dem = os.path.join(out_dir, "dem.tif")
            processing.run(
                "gdal:translate",
                {
                    "INPUT": dem,
                    "OUTPUT": out_dem,
                    "OPTIONS": ""
                }
            )
        if contours:
            out_c = os.path.join(out_dir, "contours.gpkg")
            QgsVectorFileWriter.writeAsVectorFormat(
                contours, out_c, "UTF-8", contours.crs(), "GPKG"
            )
        if boundary:
            out_b = os.path.join(out_dir, "boundary.gpkg")
            QgsVectorFileWriter.writeAsVectorFormat(
                boundary, out_b, "UTF-8", boundary.crs(), "GPKG"
            )

    def _load_boundary(self, path, target_crs):
        """Load a polygon boundary and reproject to target CRS if needed."""
        layer = QgsVectorLayer(path, "boundary_src", "ogr")
        if not layer.isValid():
            raise ValueError(f"Could not open boundary file: {path}")

        # Accept polygon / multipolygon only
        wkb = layer.wkbType()
        if not QgsWkbTypes.geometryType(wkb) == QgsWkbTypes.PolygonGeometry:
            # try dissolve of mixed; still require polygon
            raise ValueError(
                "Boundary must be a polygon layer (Polygon or MultiPolygon)."
            )

        if layer.crs() != target_crs and target_crs.isValid():
            result = processing.run(
                "native:reprojectlayer",
                {
                    "INPUT": layer,
                    "TARGET_CRS": target_crs,
                    "OUTPUT": "memory:",
                },
            )
            layer = result["OUTPUT"]

        # Dissolve to a single mask if multiple features (optional but helpful)
        if layer.featureCount() > 1:
            result = processing.run(
                "native:dissolve",
                {
                    "INPUT": layer,
                    "FIELD": [],
                    "OUTPUT": "memory:",
                },
            )
            layer = result["OUTPUT"]

        layer.setName("Boundary")
        return layer

    def _clip_to_boundary(self, contours, boundary):
        """Clip contour polylines to the boundary polygon (keeps attributes)."""
        result = processing.run(
            "native:clip",
            {
                "INPUT": contours,
                "OVERLAY": boundary,
                "OUTPUT": "memory:",
            },
        )
        clipped = result["OUTPUT"]
        if isinstance(clipped, str):
            clipped = QgsVectorLayer(clipped, "Contours", "ogr")
        return clipped


    def _export_dxf(self, dxf_path, contours=None, boundary=None, points=None):
        """Export selected layers to DXF.

        Strategy (most reliable on Windows QGIS):
        1. Write each selected layer to a temporary GeoPackage with simple attributes
        2. Prefer native:dxfexport (QGIS built-in)
        3. Fallback: QgsDxfExport API
        4. Last resort: write separate DXF files via GDAL convertformat
        """
        selected = []
        if contours is not None and contours.isValid() and contours.featureCount() > 0:
            selected.append(("Contours", contours, "ELEV"))
        if boundary is not None and boundary.isValid() and boundary.featureCount() > 0:
            selected.append(("Boundary", boundary, None))
        if points is not None and points.isValid() and points.featureCount() > 0:
            selected.append(("Points", points, "Z"))

        if not selected:
            raise ValueError("No layers selected (or available) for DXF export.")

        # Prepare clean on-disk copies (memory layers + complex attrs often break DXF)
        prepared = []
        for name, layer, elev_field in selected:
            clean = self._prepare_layer_for_dxf(layer, name, elev_field)
            prepared.append(clean)
            # Ensure visible to processing / project
            if clean.id() not in [l.id() for l in QgsProject.instance().mapLayers().values()]:
                QgsProject.instance().addMapLayer(clean, False)

        # --- 1) native:dxfexport ---
        try:
            algs = {a.id() for a in QgsApplication.processingRegistry().algorithms()}
            if "native:dxfexport" in algs:
                layer_maps = []
                for lyr in prepared:
                    # attributeIndex: field used to split DXF layers (ELEV if present)
                    attr_idx = -1
                    for i, f in enumerate(lyr.fields()):
                        if f.name().upper() in ("ELEV", "Z"):
                            attr_idx = i
                            break
                    layer_maps.append({"layer": lyr.id(), "attributeIndex": attr_idx})

                processing.run(
                    "native:dxfexport",
                    {
                        "LAYERS": layer_maps,
                        "SYMBOLOGY_MODE": 0,  # No symbology - most compatible
                        "SYMBOLOGY_SCALE": 1.0,
                        "ENCODING": "UTF-8",
                        "FORCE_2D": True,
                        "EXPORT_LABELS": False,
                        "USE_LAYER_TITLE": True,
                        "OUTPUT": dxf_path,
                    },
                )
                if os.path.isfile(dxf_path) and os.path.getsize(dxf_path) > 0:
                    self._log(f"DXF written via native:dxfexport ({len(prepared)} layer(s))")
                    return
        except Exception as e:
            self._log(f"native:dxfexport not used: {e}")

        # --- 2) QgsDxfExport API ---
        try:
            from qgis.core import QgsDxfExport, QgsMapSettings

            dxf = QgsDxfExport()
            dxf.setForce2d(True)
            dxf.setLayerTitleAsName(True)
            if hasattr(QgsDxfExport, "FlagNoMText"):
                try:
                    dxf.setFlags(QgsDxfExport.FlagNoMText)
                except (AttributeError, TypeError) as e:
                    self._log(f"DXF FlagNoMText not applied: {e}")

            dxf_layers = []
            for lyr in prepared:
                attr_idx = -1
                for i, f in enumerate(lyr.fields()):
                    if f.name().upper() in ("ELEV", "Z"):
                        attr_idx = i
                        break
                try:
                    dxf_layers.append(QgsDxfExport.DxfLayer(lyr, attr_idx))
                except TypeError:
                    dxf_layers.append(QgsDxfExport.DxfLayer(lyr))

            dxf.addLayers(dxf_layers)

            # Extent / CRS from first layer
            extent = prepared[0].extent()
            for lyr in prepared[1:]:
                extent.combineExtentWith(lyr.extent())
            ms = QgsMapSettings()
            ms.setExtent(extent)
            ms.setDestinationCrs(prepared[0].crs())
            ms.setLayers(prepared)
            dxf.setMapSettings(ms)
            if hasattr(dxf, "setDestinationCrs"):
                try:
                    dxf.setDestinationCrs(prepared[0].crs())
                except (AttributeError, TypeError) as e:
                    self._log(f"DXF setDestinationCrs not applied: {e}")

            from qgis.PyQt.QtCore import QFile, QIODevice, QTextStream
            f = QFile(dxf_path)
            if not f.open(QIODevice.WriteOnly | QIODevice.Truncate | QIODevice.Text):
                raise RuntimeError(f"Cannot open DXF for writing: {dxf_path}")
            # writeToFile signature varies by QGIS version
            ok = False
            try:
                result = dxf.writeToFile(f, "UTF-8")
                ok = bool(result) if not isinstance(result, int) else (result == 0)
            except TypeError:
                stream = QTextStream(f)
                result = dxf.writeToFile(stream, "UTF-8")
                ok = bool(result) if not isinstance(result, int) else (result == 0)
            f.close()
            if ok and os.path.isfile(dxf_path) and os.path.getsize(dxf_path) > 0:
                self._log(f"DXF written via QgsDxfExport ({len(prepared)} layer(s))")
                return
        except Exception as e:
            self._log(f"QgsDxfExport not used: {e}")

        # --- 3) Per-layer DXF via processing / OGR ---
        written = []
        base, ext = os.path.splitext(dxf_path)
        for i, lyr in enumerate(prepared):
            out = dxf_path if i == 0 else f"{base}_{lyr.name()}{ext or '.dxf'}"
            try:
                processing.run(
                    "gdal:convertformat",
                    {
                        "INPUT": lyr,
                        "OPTIONS": "-f DXF",
                        "OUTPUT": out,
                    },
                )
                if os.path.isfile(out):
                    written.append(out)
                    continue
            except Exception as e:
                self._log(f"gdal:convertformat DXF skipped for {lyr.name()}: {e}")
            # Last fallback: shapefile then hope user converts; still try writer with no attrs
            options = QgsVectorFileWriter.SaveVectorOptions()
            options.driverName = "DXF"
            options.fileEncoding = "UTF-8"
            options.onlySelectedFeatures = False
            # Strip to geometry-only layer
            geom_only = self._geometry_only_layer(lyr)
            transform_context = QgsProject.instance().transformContext()
            try:
                err = QgsVectorFileWriter.writeAsVectorFormatV3(
                    geom_only, out, transform_context, options
                )
                code = err[0] if isinstance(err, tuple) else err
            except Exception:
                code = QgsVectorFileWriter.writeAsVectorFormat(
                    geom_only, out, "UTF-8", geom_only.crs(), "DXF"
                )
            if code == QgsVectorFileWriter.NoError and os.path.isfile(out):
                written.append(out)
            else:
                self._log(f"Could not write DXF for layer {lyr.name()} (code {code})")

        if not written:
            raise RuntimeError(
                "Failed to write DXF. "
                "Try Project → Import/Export → Export Project to DXF in QGIS, "
                "or export contours as shapefile and convert in CAD."
            )
        self._log("DXF file(s) written: " + ", ".join(written))

    def _prepare_layer_for_dxf(self, layer, name, elev_field=None):
        """Copy layer to a temp GPKG with only simple numeric/text fields DXF can handle."""
        tmp = os.path.join(self.temp_dir, f"dxf_{name}.gpkg")
        # Build a memory layer with simplified attributes
        geom_type = QgsWkbTypes.displayString(QgsWkbTypes.flatType(layer.wkbType()))
        # Prefer 2D
        if "Z" in geom_type or "25D" in geom_type:
            geom_type = geom_type.replace("Z", "").replace("25D", "")
        mem = QgsVectorLayer(f"{geom_type}?crs={layer.crs().authid()}", name, "memory")
        pr = mem.dataProvider()

        new_fields = []
        keep_indices = []  # (src_index, is_elev)
        elev_src = None
        if elev_field:
            elev_src = layer.fields().indexOf(elev_field)
            if elev_src < 0:
                for f in layer.fields():
                    if f.name().upper() in ("ELEV", "Z", "ELEVATION"):
                        elev_src = layer.fields().indexOf(f.name())
                        break

        if elev_src is not None and elev_src >= 0:
            new_fields.append(QgsField("ELEV", QVariant.Double))
            keep_indices.append((elev_src, True))
        else:
            # keep first numeric field if any
            for i, f in enumerate(layer.fields()):
                if f.type() in (QVariant.Double, QVariant.Int, QVariant.LongLong):
                    new_fields.append(QgsField(f.name()[:10], QVariant.Double))
                    keep_indices.append((i, False))
                    break

        if new_fields:
            pr.addAttributes(new_fields)
            mem.updateFields()

        feats = []
        for f in layer.getFeatures():
            nf = QgsFeature(mem.fields())
            g = f.geometry()
            if g and not g.isEmpty():
                # force 2D
                try:
                    g = QgsGeometry(g)
                    if hasattr(g.get(), "dropZValue"):
                        g.get().dropZValue()
                    if hasattr(g.get(), "dropMValue"):
                        g.get().dropMValue()
                except (AttributeError, TypeError) as e:
                    self._log(f"Could not drop Z/M values: {e}")
                nf.setGeometry(g)
            attrs = []
            for src_i, _ in keep_indices:
                try:
                    attrs.append(float(f.attributes()[src_i]))
                except (TypeError, ValueError, IndexError):
                    attrs.append(None)
            if attrs:
                nf.setAttributes(attrs)
            feats.append(nf)
        if feats:
            pr.addFeatures(feats)
        mem.updateExtents()

        # Write to disk GPKG for stability
        QgsVectorFileWriter.writeAsVectorFormat(mem, tmp, "UTF-8", mem.crs(), "GPKG")
        disk = QgsVectorLayer(tmp, name, "ogr")
        if not disk.isValid():
            return mem
        return disk

    def _geometry_only_layer(self, layer):
        """Return a memory copy with no attributes (maximum DXF compatibility)."""
        geom_type = QgsWkbTypes.displayString(QgsWkbTypes.flatType(layer.wkbType()))
        mem = QgsVectorLayer(f"{geom_type}?crs={layer.crs().authid()}", layer.name(), "memory")
        pr = mem.dataProvider()
        feats = []
        for f in layer.getFeatures():
            nf = QgsFeature()
            g = f.geometry()
            if g and not g.isEmpty():
                try:
                    g = QgsGeometry(g)
                    if hasattr(g.get(), "dropZValue"):
                        g.get().dropZValue()
                    if hasattr(g.get(), "dropMValue"):
                        g.get().dropMValue()
                except (AttributeError, TypeError) as e:
                    self._log(f"Could not drop Z/M values: {e}")
                nf.setGeometry(g)
            feats.append(nf)
        pr.addFeatures(feats)
        mem.updateExtents()
        return mem
